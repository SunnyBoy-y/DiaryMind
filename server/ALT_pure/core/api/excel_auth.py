import os
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl import Workbook
import time

# Excel文件路径
BASE_DIR = Path(__file__).resolve().parent.parent.parent.parent.parent
EXCEL_PATH = BASE_DIR / "server" / "users.xlsx"
LOCK_PATH = EXCEL_PATH.with_suffix(".lock")

class FileLockTimeout(Exception):
    pass

def _acquire_lock(timeout: float = 5.0, poll_interval: float = 0.05):
    """
    简易文件锁，防止并发写入导致Excel损坏。
    """
    start = time.time()
    while True:
        try:
            # 使用独占创建锁文件
            fd = os.open(str(LOCK_PATH), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.close(fd)
            return
        except FileExistsError:
            if time.time() - start > timeout:
                raise FileLockTimeout("Lock wait timeout")
            time.sleep(poll_interval)

def _release_lock():
    try:
        if os.path.exists(LOCK_PATH):
            os.remove(LOCK_PATH)
    except Exception:
        pass

def init_db():
    """初始化Excel数据库，如果文件不存在则创建"""
    if not os.path.exists(EXCEL_PATH):
        _acquire_lock()
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Users"
            # 创建表头
            ws.append(["id", "username", "email", "password", "created_at"])
            wb.save(EXCEL_PATH)
            print(f"Created user database at {EXCEL_PATH}")
        finally:
            _release_lock()

def get_all_users():
    """获取所有用户"""
    init_db()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    users = []
    
    # 遍历行，跳过表头
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:  # 确保ID存在
            users.append({
                "id": row[0],
                "username": row[1],
                "email": row[2],
                "password": row[3],
                "created_at": row[4]
            })
    
    return users

def get_user_by_username(username: str):
    """通过用户名查找用户"""
    users = get_all_users()
    for user in users:
        if user["username"] == username:
            return user
    return None

def get_user_by_email(email: str):
    """通过邮箱查找用户"""
    users = get_all_users()
    for user in users:
        if user["email"] == email:
            return user
    return None

def create_user(username, email, password_hash):
    """创建新用户"""
    init_db()
    _acquire_lock()
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        
        # 获取下一个ID
        max_row = ws.max_row
        new_id = 1
        if max_row > 1:
            # 尝试从最后一行获取ID
            last_id = ws.cell(row=max_row, column=1).value
            if isinstance(last_id, int):
                new_id = last_id + 1
            else:
                # 如果最后一行ID无效，遍历所有行找到最大ID
                ids = [row[0] for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if isinstance(row[0], int)]
                if ids:
                    new_id = max(ids) + 1

        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws.append([new_id, username, email, password_hash, created_at])
        wb.save(EXCEL_PATH)
        
        return {
            "id": new_id,
            "username": username,
            "email": email,
            "created_at": created_at
        }
    finally:
        _release_lock()
